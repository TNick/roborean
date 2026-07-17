"""XLSX document driver using openpyxl."""

import io
from collections.abc import Mapping
from dataclasses import dataclass, field
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from pydantic import TypeAdapter
from roborean_documents_base.capabilities import assert_op_allowed
from roborean_documents_base.resolve_values import public_literal_value
from roborean_documents_base.template_store import DocumentTemplateStore
from roborean_spec import (
    DocumentDriverManifest,
    DocumentOperation,
    DocumentPreview,
    TemplateManifest,
    WorkspaceValue,
)


@dataclass
class XlsxSession:
    """Open workbook session.

    Attributes:
        document_id: Document definition identifier for this session.
        driver_id: Driver id that owns the session.
        workbook: openpyxl workbook being edited.
        ops_applied: Serialized operations applied so far.
    """

    document_id: str
    driver_id: str
    workbook: Any
    ops_applied: list[dict[str, Any]] = field(default_factory=list)


class XlsxDocumentDriver:
    """Spreadsheet driver with sheet.* operations.

    Attributes:
        driver_id: Stable driver identifier.
        manifest: Driver capability and media-type manifest.

        _template: Optional ``.xlsx`` template bytes.
        _manifest: Template sidecar manifest from ``load_template``.
    """

    driver_id: str = "roborean.xlsx"
    manifest: DocumentDriverManifest = DocumentDriverManifest(
        driverId="roborean.xlsx",
        version="0.3.0",
        irFamily="sheet",
        capabilities=[
            "set_metadata",
            "replace_named_value",
            "sheet.set_cell",
            "sheet.set_formula",
            "sheet.ensure_sheet",
            "finalize",
        ],
        supportsPreview=True,
        supportsBrowserExecution=False,
        supportsDiff=True,
        requiresBackend=True,
        templateMediaTypes=[
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ],
    )

    _template: bytes | None
    _manifest: TemplateManifest | None

    def __init__(self) -> None:
        """Initialize empty template state."""
        self._template = None
        self._manifest = None

    def load_template(
        self,
        template_ref: str,
        *,
        store: DocumentTemplateStore,
        manifest: TemplateManifest,
    ) -> None:
        """Load ``.xlsx`` template bytes.

        Args:
            template_ref: Template identifier within the project package.
            store: Template store used to resolve template bytes.
            manifest: Validated template sidecar manifest.
        """
        self._template = store.load_bytes(template_ref)
        self._manifest = manifest

    def begin_session(
        self,
        workspace: Any,
        metadata: Mapping[str, Any],
    ) -> XlsxSession:
        """Open a workbook from the template or a blank book.

        Args:
            workspace: Current workspace snapshot (unused for XLSX).
            metadata: Session metadata such as ``documentId``.

        Returns:
            Open session bound to an openpyxl workbook.
        """
        if self._template:
            workbook = load_workbook(io.BytesIO(self._template))
        else:
            workbook = Workbook()

        return XlsxSession(
            document_id=str(metadata.get("documentId", "")),
            driver_id=self.driver_id,
            workbook=workbook,
        )

    def apply_operation(
        self, session: XlsxSession, op: DocumentOperation
    ) -> None:
        """Apply one sheet operation.

        Args:
            session: Open workbook session that receives the operation.
            op: Typed document operation to apply.

        Raises:
            UnsupportedOperationError: When the op is outside capabilities.
            DriverError: When a cell value is not a public literal.
        """
        assert_op_allowed(self.manifest, op)
        data = op.model_dump(mode="python", by_alias=True)
        session.ops_applied.append(op.model_dump(mode="json", by_alias=True))

        # Dispatch sheet mutations onto the openpyxl workbook.
        if op.op == "sheet.ensure_sheet":
            name = str(data["name"])
            if name not in session.workbook.sheetnames:
                session.workbook.create_sheet(name)
            return

        if op.op == "sheet.set_cell":
            sheet = session.workbook[str(data["sheet"])]
            value = TypeAdapter(WorkspaceValue).validate_python(data["value"])
            sheet[str(data["cell"])] = public_literal_value(value)
            return

        if op.op == "sheet.set_formula":
            sheet = session.workbook[str(data["sheet"])]
            formula = str(data["formula"])
            if not formula.startswith("="):
                formula = "=" + formula
            sheet[str(data["cell"])] = formula

    def finalize(self, session: XlsxSession) -> None:
        """No-op finalize.

        Args:
            session: Open session about to be serialized.
        """
        return

    def serialize(self, session: XlsxSession) -> bytes:
        """Return ``.xlsx`` bytes.

        Args:
            session: Finalized session to serialize.

        Returns:
            Binary Office Open XML spreadsheet payload.
        """
        buffer = io.BytesIO()
        session.workbook.save(buffer)
        return buffer.getvalue()

    def preview(self, session: XlsxSession) -> DocumentPreview:
        """Build a simplified HTML table preview.

        Args:
            session: Finalized session to preview.

        Returns:
            HTML preview with truncated sheet tables.
        """
        max_rows = 50
        parts = ['<div class="roborean-xlsx">']

        # Emit a small HTML table per sheet for browser consumers.
        for name in session.workbook.sheetnames:
            sheet = session.workbook[name]
            parts.append(f"<h3>{name}</h3><table>")
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index >= max_rows:
                    break
                cells = "".join(
                    f"<td>{'' if c is None else c}</td>" for c in row
                )
                parts.append(f"<tr>{cells}</tr>")
            parts.append("</table>")
        parts.append("</div>")

        return DocumentPreview(
            documentId=session.document_id,
            mode="html",
            body="".join(parts),
            warnings=[],
            generatedAt=datetime.now(UTC).isoformat(),
            renderer={
                "package": "roborean-documents-xlsx",
                "version": "0.3.0",
            },
        )


def create_driver() -> XlsxDocumentDriver:
    """Entry-point factory.

    Returns:
        New ``XlsxDocumentDriver`` instance.
    """
    return XlsxDocumentDriver()


def xlsx_semantic_equal(a: bytes, b: bytes) -> bool:
    """Compare sheet values/formulas; ignore workbook metadata.

    Args:
        a: First serialized ``.xlsx`` artifact.
        b: Second serialized ``.xlsx`` artifact.

    Returns:
        True when sheet names and cell values/formulas match.
    """
    left = load_workbook(io.BytesIO(a), data_only=False)
    right = load_workbook(io.BytesIO(b), data_only=False)

    if left.sheetnames != right.sheetnames:
        return False

    for name in left.sheetnames:
        lsheet = left[name]
        rsheet = right[name]
        lvals = [[cell.value for cell in row] for row in lsheet.iter_rows()]
        rvals = [[cell.value for cell in row] for row in rsheet.iter_rows()]
        if lvals != rvals:
            return False

    return True
